import base64
from datetime import date, datetime
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError
from odoo.tools.misc import xlsxwriter

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled with user-facing error
    load_workbook = None


class C18OkrExcelImportWizard(models.TransientModel):
    _name = "c18.okr.excel.import.wizard"
    _description = "Import OKR from Excel"

    file_data = fields.Binary(required=True, string="Excel File")
    file_name = fields.Char(required=True, string="Filename")
    sheet_name = fields.Char(
        string="Sheet Name",
        help="Optional. Leave empty to use the first sheet.",
    )
    create_cycle_if_missing = fields.Boolean(default=False)
    create_perspective_if_missing = fields.Boolean(default=True)
    update_existing_kpi = fields.Boolean(default=True)
    activate_objective = fields.Boolean(default=True)

    _expected_columns = [
        "cycle_name",
        "objective_name",
        "perspective_name",
        "objective_scope",
        "hierarchy_level",
        "department_name",
        "employee_name",
        "objective_owner_login",
        "progress_method",
        "alignment_weight",
        "parent_objective_name",
        "kpi_name",
        "kr_type",
        "start_value",
        "target_value",
        "target_mode",
        "kpi_weight",
        "is_lower_better",
        "kpi_owner_login",
        "data_source",
        "auto_checkin_enabled",
        "cycle_start_date",
        "cycle_end_date",
        "cycle_review_date",
    ]

    def _normalize_header(self, value):
        return str(value or "").strip().lower().replace(" ", "_")

    def _to_bool(self, value, default=False):
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        as_text = str(value).strip().lower()
        return as_text in {"1", "true", "yes", "y", "x"}

    def _to_float(self, value, default=0.0):
        if value in (None, ""):
            return default
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(_("Cannot convert '%s' to number.") % value) from exc

    def _to_date(self, value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        raise ValidationError(_("Invalid date format: %s") % value)

    def _validate_choice(self, field_name, value, fallback):
        if not value:
            return fallback
        allowed = {key for key, _label in self.env["c18.okr.kpi"]._fields[field_name].selection}
        if value not in allowed:
            raise ValidationError(_("Invalid value '%(value)s' for %(field)s.") % {"value": value, "field": field_name})
        return value

    def _resolve_user(self, login_or_email):
        if not login_or_email:
            return self.env.user
        user = self.env["res.users"].search(
            ["|", ("login", "=", login_or_email), ("email", "=", login_or_email)],
            limit=1,
        )
        if not user:
            raise ValidationError(_("User '%s' was not found.") % login_or_email)
        return user

    def _resolve_department(self, department_name):
        if not department_name:
            return False
        department = self.env["hr.department"].search([("name", "=", department_name)], limit=1)
        if not department:
            raise ValidationError(_("Department '%s' was not found.") % department_name)
        return department

    def _resolve_employee(self, employee_name):
        if not employee_name:
            return False
        employee = self.env["hr.employee"].search([("name", "=", employee_name)], limit=1)
        if not employee:
            raise ValidationError(_("Employee '%s' was not found.") % employee_name)
        return employee

    def _prepare_rows(self):
        self.ensure_one()
        if load_workbook is None:
            raise UserError(
                _("Python package 'openpyxl' is required for Excel import. Please install it in the Odoo environment.")
            )
        if not self.file_data:
            raise UserError(_("Please upload an Excel file first."))

        workbook = load_workbook(filename=BytesIO(base64.b64decode(self.file_data)), data_only=True, read_only=True)
        worksheet = workbook[self.sheet_name] if self.sheet_name else workbook.worksheets[0]
        raw_rows = list(worksheet.iter_rows(values_only=True))
        if not raw_rows:
            raise UserError(_("The selected worksheet is empty."))

        headers = [self._normalize_header(value) for value in raw_rows[0]]
        missing = [column for column in ("cycle_name", "objective_name", "perspective_name", "kpi_name", "target_value") if column not in headers]
        if missing:
            raise UserError(_("Missing required columns: %s") % ", ".join(missing))

        rows = []
        for row_number, row_values in enumerate(raw_rows[1:], start=2):
            if not any(cell not in (None, "") for cell in row_values):
                continue
            row_data = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row_data[header] = row_values[index] if index < len(row_values) else False
            row_data["__row_number__"] = row_number
            rows.append(row_data)

        if not rows:
            raise UserError(_("No data rows found in the worksheet."))
        return rows

    def _get_or_create_cycle(self, row_data, cycle_cache):
        cycle_name = str(row_data.get("cycle_name") or "").strip()
        if not cycle_name:
            raise ValidationError(_("cycle_name is required."))
        cache_key = cycle_name.lower()
        if cache_key in cycle_cache:
            return cycle_cache[cache_key]

        cycle = self.env["c18.okr.cycle"].search([("name", "=", cycle_name)], limit=1)
        if cycle:
            cycle_cache[cache_key] = cycle
            return cycle

        if not self.create_cycle_if_missing:
            raise ValidationError(_("Cycle '%s' not found. Enable 'Create cycle if missing' or create it first.") % cycle_name)

        start_date = self._to_date(row_data.get("cycle_start_date")) or fields.Date.today()
        end_date = self._to_date(row_data.get("cycle_end_date")) or start_date
        if end_date < start_date:
            raise ValidationError(_("cycle_end_date cannot be before cycle_start_date for cycle '%s'.") % cycle_name)

        cycle = self.env["c18.okr.cycle"].create({
            "name": cycle_name,
            "period_type": "custom",
            "date_start": start_date,
            "date_end": end_date,
            "review_date": self._to_date(row_data.get("cycle_review_date")),
        })
        cycle_cache[cache_key] = cycle
        return cycle

    def _get_or_create_perspective(self, perspective_name, perspective_cache):
        key = perspective_name.lower()
        if key in perspective_cache:
            return perspective_cache[key]

        perspective = self.env["c18.bsc.perspective"].search([("name", "=", perspective_name)], limit=1)
        if perspective:
            perspective_cache[key] = perspective
            return perspective

        if not self.create_perspective_if_missing:
            raise ValidationError(_("Perspective '%s' not found.") % perspective_name)

        perspective = self.env["c18.bsc.perspective"].create({"name": perspective_name})
        perspective_cache[key] = perspective
        return perspective

    def action_import_excel(self):
        self.ensure_one()
        rows = self._prepare_rows()

        objective_scope_allowed = {key for key, _label in self.env["c18.okr.objective"]._fields["objective_scope"].selection}
        hierarchy_allowed = {key for key, _label in self.env["c18.okr.objective"]._fields["hierarchy_level"].selection}
        progress_method_allowed = {key for key, _label in self.env["c18.okr.objective"]._fields["progress_method"].selection}
        data_source_allowed = {key for key, _label in self.env["c18.okr.kpi"]._fields["data_source"].selection}

        cycle_cache = {}
        perspective_cache = {}
        objective_cache = {}
        created_objectives = 0
        created_kpis = 0
        updated_kpis = 0

        for row_data in rows:
            try:
                cycle = self._get_or_create_cycle(row_data, cycle_cache)
                perspective_name = str(row_data.get("perspective_name") or "").strip()
                if not perspective_name:
                    raise ValidationError(_("perspective_name is required."))
                perspective = self._get_or_create_perspective(perspective_name, perspective_cache)

                objective_name = str(row_data.get("objective_name") or "").strip()
                if not objective_name:
                    raise ValidationError(_("objective_name is required."))

                objective_scope = str(row_data.get("objective_scope") or "department").strip().lower()
                if objective_scope not in objective_scope_allowed:
                    objective_scope = "department"

                hierarchy_level = str(row_data.get("hierarchy_level") or "").strip().lower()
                hierarchy_level = hierarchy_level if hierarchy_level in hierarchy_allowed else False

                progress_method = str(row_data.get("progress_method") or "weighted").strip().lower()
                if progress_method not in progress_method_allowed:
                    progress_method = "weighted"

                department = self._resolve_department(str(row_data.get("department_name") or "").strip())
                employee = self._resolve_employee(str(row_data.get("employee_name") or "").strip())
                objective_owner = self._resolve_user(str(row_data.get("objective_owner_login") or "").strip())
                alignment_weight = self._to_float(row_data.get("alignment_weight"), default=1.0)

                parent_objective = False
                parent_name = str(row_data.get("parent_objective_name") or "").strip()
                if parent_name:
                    parent_key = (cycle.id, parent_name.lower())
                    parent_objective = objective_cache.get(parent_key) or self.env["c18.okr.objective"].search(
                        [("cycle_id", "=", cycle.id), ("name", "=", parent_name)],
                        limit=1,
                    )
                    if not parent_objective:
                        raise ValidationError(
                            _("Parent objective '%(parent)s' was not found in cycle '%(cycle)s'.")
                            % {"parent": parent_name, "cycle": cycle.name}
                        )

                objective_key = (cycle.id, objective_name.lower())
                objective = objective_cache.get(objective_key) or self.env["c18.okr.objective"].search(
                    [("cycle_id", "=", cycle.id), ("name", "=", objective_name)],
                    limit=1,
                )
                if not objective:
                    objective = self.env["c18.okr.objective"].create({
                        "name": objective_name,
                        "cycle_id": cycle.id,
                        "objective_scope": objective_scope,
                        "hierarchy_level": hierarchy_level,
                        "department_id": department.id if department else False,
                        "employee_id": employee.id if employee else False,
                        "perspective_id": perspective.id,
                        "parent_objective_id": parent_objective.id if parent_objective else False,
                        "alignment_weight": alignment_weight,
                        "progress_method": progress_method,
                        "owner_id": objective_owner.id,
                    })
                    if self.activate_objective and objective.state == "draft":
                        objective.action_activate()
                    created_objectives += 1
                objective_cache[objective_key] = objective

                kpi_name = str(row_data.get("kpi_name") or "").strip()
                if not kpi_name:
                    raise ValidationError(_("kpi_name is required."))

                kr_type = str(row_data.get("kr_type") or "numeric").strip().lower()
                kr_type = self._validate_choice("kr_type", kr_type, "numeric")

                target_mode = str(row_data.get("target_mode") or "overall").strip().lower()
                target_mode = self._validate_choice("target_mode", target_mode, "overall")

                data_source = str(row_data.get("data_source") or "manual").strip().lower()
                data_source = data_source if data_source in data_source_allowed else "manual"

                kpi_values = {
                    "name": kpi_name,
                    "objective_id": objective.id,
                    "kr_type": kr_type,
                    "start_value": self._to_float(row_data.get("start_value"), default=0.0),
                    "target_value": self._to_float(row_data.get("target_value"), default=0.0),
                    "target_mode": target_mode,
                    "weight": self._to_float(row_data.get("kpi_weight"), default=1.0),
                    "is_lower_better": self._to_bool(row_data.get("is_lower_better"), default=False),
                    "owner_id": self._resolve_user(str(row_data.get("kpi_owner_login") or "").strip()).id,
                    "data_source": data_source,
                    "auto_checkin_enabled": self._to_bool(row_data.get("auto_checkin_enabled"), default=False),
                }

                existing_kpi = self.env["c18.okr.kpi"].search(
                    [("objective_id", "=", objective.id), ("name", "=", kpi_name)],
                    limit=1,
                )
                if existing_kpi:
                    if self.update_existing_kpi:
                        existing_kpi.write(kpi_values)
                        updated_kpis += 1
                else:
                    self.env["c18.okr.kpi"].create(kpi_values)
                    created_kpis += 1

            except ValidationError as exc:
                raise UserError(_("Import failed at row %(row)s: %(error)s") % {"row": row_data["__row_number__"], "error": exc}) from exc

        message = _(
            "OKR import completed. Created Objectives: %(obj)s, Created KPIs: %(kpi)s, Updated KPIs: %(upd)s."
        ) % {"obj": created_objectives, "kpi": created_kpis, "upd": updated_kpis}

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import Completed"),
                "message": message,
                "type": "success",
                "sticky": False,
            },
        }

    def action_download_template(self):
        self.ensure_one()
        stream = BytesIO()
        workbook = xlsxwriter.Workbook(stream, {"in_memory": True})
        worksheet = workbook.add_worksheet("OKR Import Template")
        header_format = workbook.add_format({"bold": True, "bg_color": "#DDEBF7"})

        for column, title in enumerate(self._expected_columns):
            worksheet.write(0, column, title, header_format)
            worksheet.set_column(column, column, 22)

        sample_row = [
            "Q2 2026",
            "Improve Workforce Readiness",
            "Learning & Growth",
            "department",
            "manager",
            "Administration",
            "",
            "c18.hr.manager",
            "weighted",
            1.0,
            "",
            "Training Plan Completion",
            "percentage",
            0,
            100,
            "overall",
            1.0,
            False,
            "c18.hr.manager",
            "manual",
            False,
            "2026-04-01",
            "2026-06-30",
            "2026-07-10",
        ]
        for column, value in enumerate(sample_row):
            worksheet.write(1, column, value)

        workbook.close()
        file_bytes = stream.getvalue()

        attachment = self.env["ir.attachment"].create({
            "name": "c18_okr_import_template.xlsx",
            "type": "binary",
            "datas": base64.b64encode(file_bytes),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id": self.id,
        })
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
